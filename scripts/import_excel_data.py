#!/usr/bin/env python3
"""Excel data migration: Final.xlsx + Lead Tracking List.xlsx → Railway Dev PostgreSQL.

Phases:
  1. Import Final.xlsx → customers + properties + jobs
  2. Import Lead Tracking List.xlsx → leads + customers + sales_entries
  3. Lead cleanup — mark converted/connected leads
  4. Verification — cross-reference campaign_responses, service_agreements, leads

Usage:
  python scripts/import_excel_data.py --dry-run   # verify first
  python scripts/import_excel_data.py              # commit to dev
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DB_URL = (
    "postgresql://postgres:dzfRJOuVfmNHZWQmrNFhwJTrXwTEYOSu"
    "@shortline.proxy.rlwy.net:47551/railway"
)

EXCEL_DIR = Path(__file__).resolve().parent.parent / "Viktor_excel" / "migration"
FINAL_XLSX = EXCEL_DIR / "Final.xlsx"
LEAD_TRACKING_XLSX = EXCEL_DIR / "Lead Tracking List .xlsx"

EXPECTED_ALEMBIC = "20260414_100900"

# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

JOB_TYPE_MAP = {
    "spring start up": "spring_startup",
    "new system installation": "new_installation",
    "addition & repair": "repair",
    "repairs & additions": "repair",
    "drain install": "installation",
    "sprinkler head repair": "small_repair",
    "supply pipe repair (special part)": "repair",
    "controller install": "installation",
    "valve location": "diagnostic",
    "start up & landscaping": "spring_startup",
    "start up, zone intall, and controller": "spring_startup",
    "landscpaing & irrigaiton": "installation",
    "finish system install": "installation",
    # Lead tracking job types
    "irrigation install": "new_installation",
    "sod & irrigation install": "new_installation",
    "sod install": "installation",
    "irrigation refurbish": "repair",
    "irrigation & landscaping": "installation",
    "irrigation addition": "repair",
    "irrigation refurbish or new install": "new_installation",
    "new system install": "new_installation",
    "system refurbish": "repair",
    "system update/addition": "repair",
}

READY_TO_SCHEDULE_TYPES = {
    "spring_startup",
    "small_repair",
    "diagnostic",
    "service_call",
}

LEAD_SITUATION_MAP = {
    "new_installation": "new_system",
    "installation": "new_system",
    "repair": "upgrade",
    "small_repair": "repair",
    "diagnostic": "exploring",
    "service_call": "exploring",
    "spring_startup": "seasonal_maintenance",
}


def normalize_phone(raw) -> str | None:
    """Normalize a phone value from Excel into a 10-digit DB string."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Handle Excel float (e.g. 6129900255.0)
    if "." in s:
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    # Strip all non-digit chars
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    # 11 digits starting with 1 → strip country code
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    # Valid 10-digit NANP
    if len(digits) == 10:
        return digits
    # Concatenated dual phone (≥20 digits) → take first 10
    if len(digits) >= 20:
        first = digits[:10]
        if len(first) == 10:
            return first
    return None  # malformed


def split_name(full_name: str) -> tuple[str, str]:
    """Split 'First Last' into (first_name, last_name)."""
    parts = full_name.strip().rsplit(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return parts[0].strip(), parts[0].strip()


def normalize_job_type(excel_val: str) -> str:
    if not excel_val:
        return "service_call"
    return JOB_TYPE_MAP.get(excel_val.strip().lower(), "service_call")


def job_category(jtype: str) -> str:
    return "ready_to_schedule" if jtype in READY_TO_SCHEDULE_TYPES else "requires_estimate"


def normalize_property_type(excel_val: str) -> tuple[str, bool]:
    """Returns (property_type, is_hoa)."""
    v = (excel_val or "").strip().lower()
    if v == "hoa":
        return "commercial", True
    if v == "commercial":
        return "commercial", False
    return "residential", False


def map_sales_status(excel_status: str) -> str:
    v = (excel_status or "").strip().lower()
    if "scheduled" in v and "pending" not in v:
        return "estimate_scheduled"
    return "schedule_estimate"


def parse_requested_week(raw: str) -> tuple[date | None, date | None, bool]:
    """Parse 'Requested Week Or Date(s)' → (start, end, is_tbd)."""
    if not raw:
        return None, None, True
    v = raw.strip().lower()
    if v in ("tbd", "tbh", "none", "n/a", ""):
        return None, None, True
    # Look for month/day pattern
    m = re.search(r"(\d{1,2})/(\d{1,2})", raw)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            d = date(2026, month, day)
            # Find Monday of that week
            monday = d - __import__("datetime").timedelta(days=d.weekday())
            friday = monday + __import__("datetime").timedelta(days=4)
            return monday, friday, False
        except ValueError:
            pass
    return None, None, True


def map_staffing(val: str) -> int:
    return 2 if val and "crew" in val.strip().lower() else 1


# ---------------------------------------------------------------------------
# Excel loaders
# ---------------------------------------------------------------------------


def load_final_xlsx() -> list[dict]:
    wb = openpyxl.load_workbook(FINAL_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        rows.append(
            {
                "name": str(row[0]).strip(),
                "address": str(row[1]).strip() if row[1] else "",
                "city": str(row[2]).strip() if row[2] else "",
                "phone_raw": row[3],
                "email": str(row[4]).strip() if row[4] else "",
                "single_or_crew": str(row[5]).strip() if row[5] else "",
                "extra_time": str(row[6]).strip() if row[6] else "",
                "job_type": str(row[7]).strip() if row[7] else "",
                "requested_week": str(row[8]).strip() if row[8] else "",
                "notes": str(row[9]).strip() if row[9] else "",
                "service_package": str(row[10]).strip() if row[10] else "",
                "needs_spring": str(row[11]).strip() if row[11] else "",
                "needs_summer": str(row[12]).strip() if row[12] else "",
                "needs_winter": str(row[13]).strip() if row[13] else "",
                "property_type": str(row[15]).strip() if row[15] else (
                    str(row[14]).strip() if row[14] else ""
                ),
            }
        )
    wb.close()
    return rows


def load_lead_tracking_xlsx() -> list[dict]:
    wb = openpyxl.load_workbook(LEAD_TRACKING_XLSX, read_only=True, data_only=True)
    ws = wb["Sales List"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        rows.append(
            {
                "name": str(row[0]).strip(),
                "address": str(row[1]).strip() if row[1] else "",
                "phone_raw": row[2],
                "email": str(row[3]).strip() if row[3] else "",
                "job_type": str(row[4]).strip() if row[4] else "",
                "status": str(row[5]).strip() if row[5] else "",
                "last_followup": str(row[6]).strip() if row[6] else "",
            }
        )
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Phase 1: Final.xlsx → customers + properties + jobs
# ---------------------------------------------------------------------------


def phase1_import_final(conn, rows: list[dict], report: dict) -> None:
    """Import Final.xlsx rows as customers + properties + jobs."""
    stats = report["phase1"] = {
        "existing_customers_matched": 0,
        "leads_converted": 0,
        "new_customers_created": 0,
        "properties_created": 0,
        "properties_existing": 0,
        "jobs_created": 0,
        "jobs_skipped_dup": 0,
        "tbd_jobs": 0,
        "svc_pkg_flagged": 0,
        "rows_skipped_no_phone": 0,
        "rows_skipped_malformed_phone": 0,
        "errors": [],
        "created_customers": [],
        "converted_leads": [],
        "created_jobs": [],
        "skipped_rows": [],
        "tbd_customer_names": [],
    }

    # Track phones we've already created in this run to handle duplicates within Excel
    seen_phones: dict[str, str] = {}  # phone -> customer_id

    for idx, r in enumerate(rows, start=2):
        phone = normalize_phone(r["phone_raw"])
        if phone is None:
            if r["phone_raw"] is None or str(r["phone_raw"]).strip() == "":
                stats["rows_skipped_no_phone"] += 1
            else:
                stats["rows_skipped_malformed_phone"] += 1
            stats["skipped_rows"].append(
                f"Row {idx}: {r['name']} | phone={r['phone_raw']} (no valid phone)"
            )
            continue

        # Use savepoint so individual row failures don't break the transaction
        sp = conn.begin_nested()
        try:
            first, last = split_name(r["name"])
            email = r["email"] if r["email"] and "@" in r["email"] else None

            # --- Resolve or create customer ---
            customer_id = None
            customer_action = None

            # Check if we already created this phone in a previous row
            if phone in seen_phones:
                customer_id = seen_phones[phone]
                stats["existing_customers_matched"] += 1
                customer_action = "existing"
            else:
                existing = conn.execute(
                    text("SELECT id FROM customers WHERE phone = :phone AND is_deleted = false"),
                    {"phone": phone},
                ).first()

                if existing:
                    customer_id = str(existing[0])
                    stats["existing_customers_matched"] += 1
                    customer_action = "existing"
                else:
                    # Check if they're a lead
                    lead_row = conn.execute(
                        text(
                            "SELECT id, name FROM leads WHERE phone = :phone "
                            "AND status NOT IN ('converted', 'spam', 'lost') LIMIT 1"
                        ),
                        {"phone": phone},
                    ).first()

                    # Create customer
                    result = conn.execute(
                        text(
                            "INSERT INTO customers "
                            "(first_name, last_name, phone, email, status, is_new_customer, "
                            " lead_source, sms_opt_in, email_opt_in, created_at, updated_at) "
                            "VALUES (:fn, :ln, :phone, :email, 'active', true, "
                            " 'referral', false, false, NOW(), NOW()) "
                            "RETURNING id"
                        ),
                        {"fn": first, "ln": last, "phone": phone, "email": email},
                    )
                    customer_id = str(result.scalar())

                    if lead_row:
                        conn.execute(
                            text(
                                "UPDATE leads SET status = 'converted', customer_id = :cid, "
                                "converted_at = NOW(), updated_at = NOW() WHERE id = :lid"
                            ),
                            {"cid": customer_id, "lid": lead_row[0]},
                        )
                        stats["leads_converted"] += 1
                        stats["converted_leads"].append(
                            f"{r['name']} (lead: {lead_row[1]}) -> customer {customer_id}"
                        )
                        customer_action = "converted_from_lead"
                    else:
                        stats["new_customers_created"] += 1
                        stats["created_customers"].append(
                            f"{r['name']} | {phone} | {r['city']} | {r['email']}"
                        )
                        customer_action = "new"

                seen_phones[phone] = customer_id

            # --- Resolve or create property ---
            property_id = None
            if r["address"] and r["city"]:
                prop_existing = conn.execute(
                    text(
                        "SELECT id FROM properties "
                        "WHERE customer_id = :cid AND LOWER(address) = LOWER(:addr) LIMIT 1"
                    ),
                    {"cid": customer_id, "addr": r["address"]},
                ).first()

                if prop_existing:
                    property_id = str(prop_existing[0])
                    stats["properties_existing"] += 1
                else:
                    ptype, is_hoa = normalize_property_type(r["property_type"])
                    prop_result = conn.execute(
                        text(
                            "INSERT INTO properties "
                            "(customer_id, address, city, state, property_type, "
                            " is_primary, is_hoa, system_type, created_at, updated_at) "
                            "VALUES (:cid, :addr, :city, 'MN', :ptype, "
                            " true, :hoa, 'standard', NOW(), NOW()) "
                            "RETURNING id"
                        ),
                        {
                            "cid": customer_id,
                            "addr": r["address"],
                            "city": r["city"],
                            "ptype": ptype,
                            "hoa": is_hoa,
                        },
                    )
                    property_id = str(prop_result.scalar())
                    stats["properties_created"] += 1

            # --- Create job (dedup first) ---
            jtype = normalize_job_type(r["job_type"])
            cat = job_category(jtype)

            if property_id:
                job_dup_check = conn.execute(
                    text(
                        "SELECT id FROM jobs "
                        "WHERE customer_id = :cid AND job_type = :jtype "
                        "AND property_id = :pid "
                        "AND is_deleted = false "
                        "AND status NOT IN ('completed', 'cancelled') "
                        "LIMIT 1"
                    ),
                    {"cid": customer_id, "jtype": jtype, "pid": property_id},
                ).first()
            else:
                job_dup_check = conn.execute(
                    text(
                        "SELECT id FROM jobs "
                        "WHERE customer_id = :cid AND job_type = :jtype "
                        "AND property_id IS NULL "
                        "AND is_deleted = false "
                        "AND status NOT IN ('completed', 'cancelled') "
                        "LIMIT 1"
                    ),
                    {"cid": customer_id, "jtype": jtype},
                ).first()

            if job_dup_check:
                stats["jobs_skipped_dup"] += 1
            else:
                start_dt, end_dt, is_tbd = parse_requested_week(r["requested_week"])
                staffing = map_staffing(r["single_or_crew"])

                # Build notes
                note_parts = []
                if r["notes"]:
                    note_parts.append(r["notes"])
                if r["extra_time"] and r["extra_time"].lower() not in ("no", "none", ""):
                    note_parts.append(f"Extra time: {r['extra_time']}")
                if r["requested_week"]:
                    note_parts.append(f"Requested: {r['requested_week']}")
                job_notes = "; ".join(note_parts) if note_parts else None

                conn.execute(
                    text(
                        "INSERT INTO jobs "
                        "(customer_id, property_id, job_type, category, status, "
                        " staffing_required, target_start_date, target_end_date, "
                        " notes, source, priority_level, description, "
                        " created_at, updated_at) "
                        "VALUES (:cid, :pid, :jtype, :cat, 'to_be_scheduled', "
                        " :staff, :start, :end, :notes, 'referral', 0, :desc, "
                        " NOW(), NOW())"
                    ),
                    {
                        "cid": customer_id,
                        "pid": property_id,
                        "jtype": jtype,
                        "cat": cat,
                        "staff": staffing,
                        "start": start_dt,
                        "end": end_dt,
                        "notes": job_notes,
                        "desc": f"{r['job_type']} for {r['name']}",
                    },
                )
                stats["jobs_created"] += 1
                stats["created_jobs"].append(
                    f"{r['name']} | {jtype} | {r['requested_week'] or 'TBD'} | {customer_action}"
                )

                if is_tbd:
                    stats["tbd_jobs"] += 1
                    stats["tbd_customer_names"].append(r["name"])
                    conn.execute(
                        text(
                            "UPDATE customers SET internal_notes = "
                            "COALESCE(internal_notes, '') || :note, "
                            "updated_at = NOW() WHERE id = :cid"
                        ),
                        {
                            "cid": customer_id,
                            "note": "[MIGRATION] TBD scheduling - needs outreach. ",
                        },
                    )

            # --- Flag service package edge cases ---
            if (
                r["service_package"]
                and r["service_package"].lower() == "yes"
                and customer_action == "new"
            ):
                sa_check = conn.execute(
                    text(
                        "SELECT id FROM service_agreements "
                        "WHERE customer_id = :cid AND status = 'active' LIMIT 1"
                    ),
                    {"cid": customer_id},
                ).first()
                if not sa_check:
                    stats["svc_pkg_flagged"] += 1
                    conn.execute(
                        text(
                            "UPDATE customers SET internal_notes = "
                            "COALESCE(internal_notes, '') || :note, "
                            "updated_at = NOW() WHERE id = :cid"
                        ),
                        {
                            "cid": customer_id,
                            "note": "[MIGRATION] Svc Pkg=Yes but no agreement - needs setup. ",
                        },
                    )

            sp.commit()
        except Exception as e:
            sp.rollback()
            stats["errors"].append(f"Row {idx} ({r['name']}): {e}")


# ---------------------------------------------------------------------------
# Phase 2: Lead Tracking List → leads + customers + sales_entries
# ---------------------------------------------------------------------------


def phase2_import_leads(conn, rows: list[dict], report: dict) -> None:
    """Import Lead Tracking List into leads + customers + sales pipeline."""
    stats = report["phase2"] = {
        "existing_customers_used": 0,
        "existing_leads_updated": 0,
        "new_leads_created": 0,
        "new_customers_created": 0,
        "properties_created": 0,
        "sales_entries_created": 0,
        "rows_skipped_no_phone": 0,
        "errors": [],
        "detail_rows": [],
    }

    for idx, r in enumerate(rows, start=2):
        phone = normalize_phone(r["phone_raw"])
        if phone is None:
            stats["rows_skipped_no_phone"] += 1
            continue

        sp = conn.begin_nested()
        try:
            first, last = split_name(r["name"])
            email = r["email"] if r["email"] and "@" in r["email"] else None
            jtype = normalize_job_type(r["job_type"])
            situation = LEAD_SITUATION_MAP.get(jtype, "exploring")

            # --- Resolve customer ---
            customer_id = None
            lead_id = None
            action = ""

            cust_row = conn.execute(
                text("SELECT id FROM customers WHERE phone = :phone AND is_deleted = false"),
                {"phone": phone},
            ).first()

            if cust_row:
                customer_id = str(cust_row[0])
                stats["existing_customers_used"] += 1
                action = "existing_customer"
                # Also check for existing lead to link
                lead_check = conn.execute(
                    text("SELECT id FROM leads WHERE phone = :phone LIMIT 1"),
                    {"phone": phone},
                ).first()
                if lead_check:
                    lead_id = str(lead_check[0])
            else:
                # Check if they're an existing lead
                lead_check = conn.execute(
                    text(
                        "SELECT id FROM leads WHERE phone = :phone "
                        "AND status NOT IN ('converted', 'spam', 'lost') LIMIT 1"
                    ),
                    {"phone": phone},
                ).first()

                if lead_check:
                    lead_id = str(lead_check[0])
                    stats["existing_leads_updated"] += 1
                    action = "existing_lead"
                    conn.execute(
                        text(
                            "UPDATE leads SET job_requested = :jr, "
                            "last_contacted_at = NOW(), updated_at = NOW() "
                            "WHERE id = :lid"
                        ),
                        {"jr": r["job_type"], "lid": lead_id},
                    )
                else:
                    # Create new lead
                    lead_result = conn.execute(
                        text(
                            "INSERT INTO leads "
                            "(name, phone, email, address, city, situation, status, "
                            " lead_source, source_site, sms_consent, terms_accepted, "
                            " email_marketing_consent, job_requested, "
                            " created_at, updated_at) "
                            "VALUES (:name, :phone, :email, :addr, :city, :sit, 'contacted', "
                            " 'referral', 'Viktor Excel', false, false, false, :jr, "
                            " NOW(), NOW()) "
                            "RETURNING id"
                        ),
                        {
                            "name": r["name"],
                            "phone": phone,
                            "email": email,
                            "addr": r["address"] or None,
                            "city": r["address"].split(",")[0] if "," in r["address"] else None,
                            "sit": situation,
                            "jr": r["job_type"],
                        },
                    )
                    lead_id = str(lead_result.scalar())
                    stats["new_leads_created"] += 1
                    action = "new_lead"

                # Create customer from lead data
                cust_result = conn.execute(
                    text(
                        "INSERT INTO customers "
                        "(first_name, last_name, phone, email, status, is_new_customer, "
                        " lead_source, sms_opt_in, email_opt_in, created_at, updated_at) "
                        "VALUES (:fn, :ln, :phone, :email, 'active', true, "
                        " 'referral', false, false, NOW(), NOW()) "
                        "RETURNING id"
                    ),
                    {"fn": first, "ln": last, "phone": phone, "email": email},
                )
                customer_id = str(cust_result.scalar())
                stats["new_customers_created"] += 1

                # Link lead to customer
                conn.execute(
                    text(
                        "UPDATE leads SET customer_id = :cid, moved_to = 'sales', "
                        "moved_at = NOW(), updated_at = NOW() WHERE id = :lid"
                    ),
                    {"cid": customer_id, "lid": lead_id},
                )

            # --- Create property ---
            property_id = None
            if r["address"]:
                city = ""
                addr = r["address"]
                if "," in addr:
                    parts = [p.strip() for p in addr.split(",")]
                    if len(parts) >= 2:
                        city = parts[1].replace("MN", "").replace("mn", "").strip()
                        if not city:
                            city = parts[0]
                if not city:
                    city = "Unknown"

                prop_check = conn.execute(
                    text(
                        "SELECT id FROM properties "
                        "WHERE customer_id = :cid AND LOWER(address) = LOWER(:addr) LIMIT 1"
                    ),
                    {"cid": customer_id, "addr": addr},
                ).first()

                if not prop_check:
                    prop_result = conn.execute(
                        text(
                            "INSERT INTO properties "
                            "(customer_id, address, city, state, property_type, "
                            " is_primary, is_hoa, system_type, created_at, updated_at) "
                            "VALUES (:cid, :addr, :city, 'MN', 'residential', "
                            " true, false, 'standard', NOW(), NOW()) "
                            "RETURNING id"
                        ),
                        {"cid": customer_id, "addr": addr, "city": city},
                    )
                    property_id = str(prop_result.scalar())
                    stats["properties_created"] += 1
                else:
                    property_id = str(prop_check[0])

            # --- Create SalesEntry ---
            se_dup = conn.execute(
                text(
                    "SELECT id FROM sales_entries "
                    "WHERE customer_id = :cid AND job_type = :jtype "
                    "AND status NOT IN ('closed_won', 'closed_lost') LIMIT 1"
                ),
                {"cid": customer_id, "jtype": jtype},
            ).first()

            if not se_dup:
                note_parts = []
                if r["status"]:
                    note_parts.append(f"Original status: {r['status']}")
                if r["last_followup"]:
                    note_parts.append(f"Follow-up history: {r['last_followup']}")
                se_notes = "\n".join(note_parts) if note_parts else None

                conn.execute(
                    text(
                        "INSERT INTO sales_entries "
                        "(customer_id, property_id, lead_id, job_type, status, "
                        " notes, created_at, updated_at) "
                        "VALUES (:cid, :pid, :lid, :jtype, :status, :notes, "
                        " NOW(), NOW())"
                    ),
                    {
                        "cid": customer_id,
                        "pid": property_id,
                        "lid": lead_id,
                        "jtype": jtype,
                        "status": map_sales_status(r["status"]),
                        "notes": se_notes,
                    },
                )
                stats["sales_entries_created"] += 1

            stats["detail_rows"].append(
                f"{r['name']} | {phone} | {r['job_type']} -> {jtype} | "
                f"status={r['status']} | action={action}"
            )

            sp.commit()
        except Exception as e:
            sp.rollback()
            stats["errors"].append(f"Row {idx} ({r['name']}): {e}")


# ---------------------------------------------------------------------------
# Phase 3: Lead cleanup
# ---------------------------------------------------------------------------


def phase3_lead_cleanup(conn, report: dict) -> None:
    """Mark all connected leads as converted."""
    stats = report["phase3"] = {
        "already_converted_fixed": 0,
        "phone_matched_converted": 0,
        "remaining_active_leads": 0,
        "converted_lead_names": [],
        "remaining_lead_names": [],
    }

    # Fix leads that have customer_id but aren't marked converted
    result = conn.execute(
        text(
            "UPDATE leads SET status = 'converted', converted_at = NOW(), "
            "updated_at = NOW() "
            "WHERE customer_id IS NOT NULL AND status != 'converted' "
            "RETURNING id, name"
        )
    )
    for row in result:
        stats["already_converted_fixed"] += 1
        stats["converted_lead_names"].append(f"{row[1]} (had customer_id, status fixed)")

    # Match leads by phone to existing customers
    result = conn.execute(
        text(
            "UPDATE leads SET customer_id = c.id, status = 'converted', "
            "converted_at = NOW(), updated_at = NOW() "
            "FROM customers c "
            "WHERE leads.phone = c.phone AND leads.customer_id IS NULL "
            "AND c.is_deleted = false "
            "AND leads.status NOT IN ('converted', 'spam', 'lost') "
            "RETURNING leads.id, leads.name, c.first_name || ' ' || c.last_name AS cust_name"
        )
    )
    for row in result:
        stats["phone_matched_converted"] += 1
        stats["converted_lead_names"].append(
            f"{row[1]} -> matched to customer {row[2]}"
        )

    # Count remaining active leads
    remaining = conn.execute(
        text(
            "SELECT name, phone, status FROM leads "
            "WHERE customer_id IS NULL "
            "AND status NOT IN ('converted', 'spam', 'lost') "
            "ORDER BY created_at DESC"
        )
    ).all()
    stats["remaining_active_leads"] = len(remaining)
    for row in remaining:
        stats["remaining_lead_names"].append(f"{row[0]} | {row[1]} | {row[2]}")


# ---------------------------------------------------------------------------
# Phase 4: Verification
# ---------------------------------------------------------------------------


def phase4_verify(conn, report: dict) -> None:
    """Cross-reference campaign responses, agreements, and remaining leads."""
    stats = report["phase4"] = {
        "campaign_responses_total": 0,
        "campaign_responses_with_customer": 0,
        "campaign_responses_orphan": 0,
        "campaign_orphan_details": [],
        "agreements_total": 0,
        "agreements_valid": 0,
        "agreements_invalid": 0,
        "agreements_with_stripe": 0,
        "agreements_without_stripe": 0,
        "remaining_active_leads": 0,
        "remaining_lead_details": [],
    }

    # Campaign responses
    cr_total = conn.execute(text("SELECT COUNT(*) FROM campaign_responses")).scalar()
    stats["campaign_responses_total"] = cr_total

    cr_with_cust = conn.execute(
        text("SELECT COUNT(*) FROM campaign_responses WHERE customer_id IS NOT NULL")
    ).scalar()
    stats["campaign_responses_with_customer"] = cr_with_cust

    # Try to match orphan campaign responses by phone
    orphans = conn.execute(
        text(
            "SELECT cr.id, cr.phone, cr.recipient_name "
            "FROM campaign_responses cr "
            "WHERE cr.customer_id IS NULL"
        )
    ).all()
    for orphan in orphans:
        phone = normalize_phone(orphan[1])
        matched = False
        if phone:
            cust = conn.execute(
                text("SELECT id, first_name, last_name FROM customers WHERE phone = :p"),
                {"p": phone},
            ).first()
            if cust:
                matched = True
        status = "MATCHABLE" if matched else "UNMATCHED"
        stats["campaign_orphan_details"].append(
            f"{orphan[2]} | {orphan[1]} | {status}"
        )
    stats["campaign_responses_orphan"] = len(orphans)

    # Service agreements
    sa_total = conn.execute(text("SELECT COUNT(*) FROM service_agreements")).scalar()
    stats["agreements_total"] = sa_total

    sa_valid = conn.execute(
        text(
            "SELECT COUNT(*) FROM service_agreements sa "
            "JOIN customers c ON c.id = sa.customer_id "
            "WHERE c.is_deleted = false"
        )
    ).scalar()
    stats["agreements_valid"] = sa_valid
    stats["agreements_invalid"] = sa_total - sa_valid

    sa_stripe = conn.execute(
        text(
            "SELECT COUNT(*) FROM service_agreements "
            "WHERE stripe_subscription_id IS NOT NULL"
        )
    ).scalar()
    stats["agreements_with_stripe"] = sa_stripe
    stats["agreements_without_stripe"] = sa_total - sa_stripe

    # Remaining active leads
    remaining = conn.execute(
        text(
            "SELECT name, phone, status FROM leads "
            "WHERE status NOT IN ('converted', 'spam', 'lost') "
            "AND customer_id IS NULL "
            "ORDER BY name"
        )
    ).all()
    stats["remaining_active_leads"] = len(remaining)
    for r in remaining:
        stats["remaining_lead_details"].append(f"{r[0]} | {r[1]} | {r[2]}")


# ---------------------------------------------------------------------------
# Report printer
# ---------------------------------------------------------------------------


def get_table_counts(conn) -> dict:
    counts = {}
    for table in ("customers", "leads", "jobs", "properties", "sales_entries", "service_agreements"):
        q = f"SELECT COUNT(*) FROM {table}"
        if table == "customers":
            q += " WHERE is_deleted = false"
        counts[table] = conn.execute(text(q)).scalar()
    return counts


def print_report(report: dict, dry_run: bool) -> None:
    banner = "*** DRY RUN - NO CHANGES COMMITTED ***" if dry_run else "*** CHANGES COMMITTED TO DEV ***"
    sep = "=" * 80

    print(f"\n{sep}")
    print(f"  MIGRATION REPORT — {banner}")
    print(f"{sep}")

    # Before/after
    print(f"\n--- Table Counts ---")
    print(f"  {'Table':<25} {'Before':>8} {'After':>8} {'Delta':>8}")
    for table in ("customers", "leads", "jobs", "properties", "sales_entries", "service_agreements"):
        before = report["before"].get(table, 0)
        after = report["after"].get(table, 0)
        delta = after - before
        sign = "+" if delta > 0 else ""
        print(f"  {table:<25} {before:>8} {after:>8} {sign}{delta:>7}")

    # Phase 1
    p1 = report.get("phase1", {})
    print(f"\n{sep}")
    print(f"  PHASE 1: Final.xlsx → Customers + Properties + Jobs")
    print(f"{sep}")
    print(f"  Existing customers matched:   {p1.get('existing_customers_matched', 0)}")
    print(f"  Leads converted to customers: {p1.get('leads_converted', 0)}")
    print(f"  New customers created:        {p1.get('new_customers_created', 0)}")
    print(f"  Properties created:           {p1.get('properties_created', 0)}")
    print(f"  Properties existing (reused): {p1.get('properties_existing', 0)}")
    print(f"  Jobs created:                 {p1.get('jobs_created', 0)}")
    print(f"  Jobs skipped (duplicate):     {p1.get('jobs_skipped_dup', 0)}")
    print(f"  TBD jobs (needs outreach):    {p1.get('tbd_jobs', 0)}")
    print(f"  Svc Pkg flagged (no SA):      {p1.get('svc_pkg_flagged', 0)}")
    print(f"  Rows skipped (no phone):      {p1.get('rows_skipped_no_phone', 0)}")
    print(f"  Rows skipped (bad phone):     {p1.get('rows_skipped_malformed_phone', 0)}")
    print(f"  Errors:                       {len(p1.get('errors', []))}")

    if p1.get("converted_leads"):
        print(f"\n  --- Leads Converted to Customers ({len(p1['converted_leads'])}) ---")
        for item in p1["converted_leads"]:
            print(f"    {item}")

    if p1.get("created_customers"):
        print(f"\n  --- New Customers Created ({len(p1['created_customers'])}) ---")
        for item in p1["created_customers"]:
            print(f"    {item}")

    if p1.get("tbd_customer_names"):
        print(f"\n  --- TBD Scheduling — Needs Outreach ({len(p1['tbd_customer_names'])}) ---")
        for name in p1["tbd_customer_names"]:
            print(f"    {name}")

    if p1.get("skipped_rows"):
        print(f"\n  --- Skipped Rows ({len(p1['skipped_rows'])}) ---")
        for item in p1["skipped_rows"]:
            print(f"    {item}")

    if p1.get("errors"):
        print(f"\n  --- Phase 1 Errors ---")
        for err in p1["errors"]:
            print(f"    ERROR: {err}")

    # Phase 2
    p2 = report.get("phase2", {})
    print(f"\n{sep}")
    print(f"  PHASE 2: Lead Tracking List → Leads + Sales Pipeline")
    print(f"{sep}")
    print(f"  Existing customers used:      {p2.get('existing_customers_used', 0)}")
    print(f"  Existing leads updated:       {p2.get('existing_leads_updated', 0)}")
    print(f"  New leads created:            {p2.get('new_leads_created', 0)}")
    print(f"  New customers created:        {p2.get('new_customers_created', 0)}")
    print(f"  Properties created:           {p2.get('properties_created', 0)}")
    print(f"  Sales entries created:        {p2.get('sales_entries_created', 0)}")
    print(f"  Rows skipped (no phone):      {p2.get('rows_skipped_no_phone', 0)}")
    print(f"  Errors:                       {len(p2.get('errors', []))}")

    if p2.get("detail_rows"):
        print(f"\n  --- Row-by-Row Detail ---")
        for item in p2["detail_rows"]:
            print(f"    {item}")

    if p2.get("errors"):
        print(f"\n  --- Phase 2 Errors ---")
        for err in p2["errors"]:
            print(f"    ERROR: {err}")

    # Phase 3
    p3 = report.get("phase3", {})
    print(f"\n{sep}")
    print(f"  PHASE 3: Lead Cleanup")
    print(f"{sep}")
    print(f"  Leads with customer_id fixed: {p3.get('already_converted_fixed', 0)}")
    print(f"  Leads phone-matched+converted:{p3.get('phone_matched_converted', 0)}")
    print(f"  Remaining active leads:       {p3.get('remaining_active_leads', 0)}")

    if p3.get("converted_lead_names"):
        print(f"\n  --- Converted Leads ---")
        for item in p3["converted_lead_names"]:
            print(f"    {item}")

    if p3.get("remaining_lead_names"):
        print(f"\n  --- Remaining Active Leads ---")
        for item in p3["remaining_lead_names"]:
            print(f"    {item}")

    # Phase 4
    p4 = report.get("phase4", {})
    print(f"\n{sep}")
    print(f"  PHASE 4: Verification")
    print(f"{sep}")
    print(f"  Campaign responses total:     {p4.get('campaign_responses_total', 0)}")
    print(f"  Campaign w/ customer_id:      {p4.get('campaign_responses_with_customer', 0)}")
    print(f"  Campaign orphans:             {p4.get('campaign_responses_orphan', 0)}")
    print(f"  Service agreements total:     {p4.get('agreements_total', 0)}")
    print(f"  Agreements valid:             {p4.get('agreements_valid', 0)}")
    print(f"  Agreements invalid:           {p4.get('agreements_invalid', 0)}")
    print(f"  Agreements w/ Stripe sub:     {p4.get('agreements_with_stripe', 0)}")
    print(f"  Agreements w/o Stripe sub:    {p4.get('agreements_without_stripe', 0)}")
    print(f"  Remaining active leads:       {p4.get('remaining_active_leads', 0)}")

    if p4.get("campaign_orphan_details"):
        print(f"\n  --- Campaign Response Orphans ---")
        for item in p4["campaign_orphan_details"]:
            print(f"    {item}")

    if p4.get("remaining_lead_details"):
        print(f"\n  --- Remaining Active Leads (after cleanup) ---")
        for item in p4["remaining_lead_details"]:
            print(f"    {item}")

    print(f"\n{sep}")
    print(f"  {banner}")
    print(f"{sep}\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel → Railway Dev migration")
    parser.add_argument("--dry-run", action="store_true", help="Run without committing")
    args = parser.parse_args()

    # Pre-flight
    print(f"Target: Railway Dev")
    print(f"Dry run: {args.dry_run}")
    print(f"Final.xlsx: {FINAL_XLSX} (exists={FINAL_XLSX.exists()})")
    print(f"Lead Tracking: {LEAD_TRACKING_XLSX} (exists={LEAD_TRACKING_XLSX.exists()})")

    if not FINAL_XLSX.exists() or not LEAD_TRACKING_XLSX.exists():
        print("ERROR: Excel files not found.")
        sys.exit(1)

    engine = create_engine(DB_URL, echo=False)
    report: dict = {}

    # Load Excel data
    print("\nLoading Excel files...")
    final_rows = load_final_xlsx()
    lead_rows = load_lead_tracking_xlsx()
    print(f"  Final.xlsx: {len(final_rows)} rows")
    print(f"  Lead Tracking: {len(lead_rows)} rows")

    with engine.connect() as conn:
        trans = conn.begin()
        try:
            # Alembic check
            alembic_ver = conn.execute(
                text("SELECT version_num FROM alembic_version")
            ).scalar()
            if alembic_ver != EXPECTED_ALEMBIC:
                print(
                    f"WARNING: Alembic version {alembic_ver} != expected {EXPECTED_ALEMBIC}"
                )

            # Before counts
            report["before"] = get_table_counts(conn)
            print(f"\nBefore counts: {report['before']}")

            # Phase 1
            print("\n--- Phase 1: Importing Final.xlsx ---")
            phase1_import_final(conn, final_rows, report)
            p1 = report["phase1"]
            print(
                f"  Done: {p1['new_customers_created']} new customers, "
                f"{p1['leads_converted']} leads converted, "
                f"{p1['jobs_created']} jobs created"
            )

            # Phase 2
            print("\n--- Phase 2: Importing Lead Tracking List ---")
            phase2_import_leads(conn, lead_rows, report)
            p2 = report["phase2"]
            print(
                f"  Done: {p2['new_customers_created']} new customers, "
                f"{p2['sales_entries_created']} sales entries"
            )

            # Phase 3
            print("\n--- Phase 3: Lead Cleanup ---")
            phase3_lead_cleanup(conn, report)
            p3 = report["phase3"]
            print(
                f"  Done: {p3['already_converted_fixed'] + p3['phone_matched_converted']} "
                f"leads converted, {p3['remaining_active_leads']} remaining"
            )

            # After counts (within same transaction)
            report["after"] = get_table_counts(conn)

            # Phase 4
            print("\n--- Phase 4: Verification ---")
            phase4_verify(conn, report)

            if args.dry_run:
                trans.rollback()
            else:
                trans.commit()

        except Exception:
            trans.rollback()
            raise

    engine.dispose()

    # Print full report
    print_report(report, args.dry_run)


if __name__ == "__main__":
    main()
